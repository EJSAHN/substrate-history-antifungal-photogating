from __future__ import annotations

from dataclasses import dataclass
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class SheetSpec:
    """One sheet to be written to the output workbook."""

    name: str
    df: pd.DataFrame
    source_label: str


def _write_df(ws: Worksheet, df: pd.DataFrame) -> None:
    df = df.copy()

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        ws.append(row)
        if r_idx == 1:
            for c_idx in range(1, len(row) + 1):
                cell = ws.cell(row=1, column=c_idx)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col in ws.columns:
        col_letter = col[0].column_letter
        values = ["" if c.value is None else str(c.value) for c in col]
        max_len = max((len(v) for v in values), default=10)
        ws.column_dimensions[col_letter].width = max(8, min(45, max_len + 2))


def build_workbook(sheets: Iterable[SheetSpec]) -> Workbook:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for spec in sheets:
        ws = wb.create_sheet(title=spec.name[:31])
        _write_df(ws, spec.df)

    return wb
